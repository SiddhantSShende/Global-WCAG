"""XLSX report builder (openpyxl): Summary + Checklist + Findings sheets.

Conditional formatting colours each criterion by status; freeze panes and
autofilter make it a working audit workbook.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from ..models import Status
from . import common

HEADER_FILL = PatternFill("solid", fgColor="0F3D57")
HEADER_FONT = Font(bold=True, color="FFFFFF")

_STATUS_FILL = {
    "Fail": PatternFill("solid", fgColor="F4CCCC"),
    "Partial": PatternFill("solid", fgColor="FCE5CD"),
    "Pass": PatternFill("solid", fgColor="D9EAD3"),
    "Needs Manual Review": PatternFill("solid", fgColor="FFF2CC"),
    "Not Applicable": PatternFill("solid", fgColor="EFEFEF"),
}


def _header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _summary_sheet(wb, combo, meta) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Accessibility Compliance Audit"
    ws["A1"].font = Font(bold=True, size=14, color="0F3D57")
    rows = [
        ("Standard", f"WCAG {combo['version']} — Level {combo['level']} (cumulative)"),
        ("Target", meta["target_ref"]),
        ("Scan date", meta["scan_date"][:19].replace("T", " ")),
        ("Pages scanned", meta.get("pages_scanned", 0)),
        ("Templates", meta.get("templates", 0)),
        ("", ""),
        ("Criteria in scope", combo["coverage"]["total_in_scope"]),
        ("Active (excl. obsolete)", combo["coverage"]["active_in_scope"]),
        ("", ""),
    ]
    for label, val in combo["counts"].items():
        rows.append((label, val))
    rows.append(("", ""))
    for label, val in combo["severity"].items():
        rows.append((f"Severity: {label}", val))
    rs = combo.get("review_stats", {})
    rows.append(("", ""))
    rows.append(("Manually verified", rs.get("manually_verified", 0)))
    rows.append(("Open — needs manual review", rs.get("open_review", 0)))
    for r, (k, v) in enumerate(rows, start=3):
        ws.cell(row=r, column=1, value=k).font = Font(bold=bool(k))
        ws.cell(row=r, column=2, value=v)
    _autosize(ws, {1: 32, 2: 60})


def _checklist_sheet(wb, combo) -> None:
    ws = wb.create_sheet("Checklist")
    headers = ["SC", "Name", "Level", "Principle", "Guideline", "Status", "Note"]
    _header(ws, headers)
    for r, row in enumerate(sorted(
            combo["rows"], key=lambda x: [int(n) for n in x["crit"]["num"].split(".")]), start=2):
        c = row["crit"]
        ws.cell(row=r, column=1, value=c["num"])
        ws.cell(row=r, column=2, value=c["name"])
        ws.cell(row=r, column=3, value=c["level"])
        ws.cell(row=r, column=4, value=c["principle"])
        ws.cell(row=r, column=5, value=c["guideline"])
        ws.cell(row=r, column=6, value=common.status_label(row["status"]))
        ws.cell(row=r, column=7, value=row["note"]).alignment = Alignment(wrap_text=True)
    last = ws.max_row
    rng = f"F2:F{last}"
    for label, fill in _STATUS_FILL.items():
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="equal", formula=[f'"{label}"'], fill=fill))
    _autosize(ws, {1: 10, 2: 40, 3: 8, 4: 16, 5: 22, 6: 22, 7: 60})


def _findings_sheet(wb, combo) -> None:
    ws = wb.create_sheet("Findings")
    headers = ["SC", "Name", "Level", "Impact", "Detected by", "Confidence",
               "Occurrences", "Locations", "Rule", "Description", "Remediation",
               "Evidence key"]
    _header(ws, headers)
    for r, f in enumerate(combo["issues"], start=2):
        locs = "; ".join(f"{loc.ref} ({loc.count})" for loc in f.locations)
        vals = [f.sc_num, f.sc_name, f.level, f.impact.value,
                ", ".join(f.engines_agreeing) or f.engine, f.confidence.value,
                f.occurrences, locs, f.engine_rule_id, f.description,
                f.remediation, f.screenshot_key or ""]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col, value=v)
            if col in (8, 10, 11):
                cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=4).fill = _STATUS_FILL["Fail"]
    _autosize(ws, {1: 10, 2: 28, 3: 8, 4: 12, 5: 22, 6: 12, 7: 12,
                   8: 40, 9: 20, 10: 45, 11: 50, 12: 40})


def _legend_sheet(wb) -> None:
    ws = wb.create_sheet("Legend")
    ws["A1"] = "Status legend"
    ws["A1"].font = Font(bold=True)
    for r, (label, fill) in enumerate(_STATUS_FILL.items(), start=2):
        ws.cell(row=r, column=1, value=label).fill = fill
    ws.cell(row=len(_STATUS_FILL) + 3, column=1,
            value="Pass = machine-decidable & clean, or human-signed-off. "
                  "Needs Manual Review = we could not prove it automatically.")
    _autosize(ws, {1: 90})


def build_xlsx(combo: dict, meta: dict, out_path: str) -> str:
    wb = Workbook()
    _summary_sheet(wb, combo, meta)
    _checklist_sheet(wb, combo)
    _findings_sheet(wb, combo)
    _legend_sheet(wb)
    wb.save(out_path)
    return out_path
